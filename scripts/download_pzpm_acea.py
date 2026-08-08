#!/usr/bin/env python3
"""Download ACEA monthly new-car registrations (all European markets) via PZPM.

ACEA's monthly press release is published as an .xlsx by PZPM (the Polish
automotive industry association), which — unlike ACEA's own PDF — is a clean,
structured workbook. The ``Market (monthly)`` sheet lists every European market
(EU27 + EFTA + UK) with the month's registrations split by power source
(BEV / PHEV / HEV / other / petrol / diesel) and a total.

VRS uses this as a *second tier*: the eight countries with their own open
national feeds keep those (with brand/model detail); every other European market
is filled from ACEA at country-total + powertrain level. ACEA data is © ACEA and
must be credited as the source wherever shown.

Scrapes the PZPM per-year listing pages for the monthly .xlsx links, downloads
new months incrementally, and writes
``data/Europe/acea_market.csv`` (year, month, country, total, bev, phev, hev,
other, petrol, diesel).
"""
from __future__ import annotations

import argparse
import csv
import io
import re
import sys
import time
import urllib.request
import zipfile
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "data" / "Europe"
OUT_CSV = OUT_DIR / "acea_market.csv"

BASE = "https://www.pzpm.org.pl"
ROOT = "/en/Europe/EUROPE-Registrations-of-vehicles/PASSENGER-CARS"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) VRS/1.0"

_FULL = ["january", "february", "march", "april", "may", "june", "july",
         "august", "september", "october", "november", "december"]
# The filenames mix full and 3-letter month names (e.g. "Jan_2024", "May_2024").
MONTHS = {}
for _i, _name in enumerate(_FULL, start=1):
    MONTHS[_name] = _i
    MONTHS[_name[:3]] = _i
MONTHS["sept"] = 9

# Year-archive slugs are irregular on the site (Year-20222, Year-20252, …); we
# both discover them from the pages and probe a fixed set.
_YEAR_SLUGS = ["Year-2020", "Year-2021", "Year-2022", "Year-20222", "Year-2023",
               "Year-2024", "Year-20242", "Year-2025", "Year-20252", "Year-2026"]

# Aggregate rows in the sheet that are not individual countries.
AGGREGATES = {"EUROPEAN UNION", "EFTA", "EU + EFTA + UK", "EU15", "EU12",
              "EU27", "EU + EFTA", "EU14 + EFTA"}

# Value columns (0-indexed) per power source. Each source spans 3 cols: this
# year / last year / % change. We read BOTH years, so each monthly file yields
# the current month AND the year-earlier month (doubling the history for free).
CUR_COLS = {"bev": 1, "phev": 4, "hev": 7, "other": 10, "petrol": 13, "diesel": 16, "total": 19}
PRIOR_COLS = {k: v + 1 for k, v in CUR_COLS.items()}


def _get(url: str, retries: int = 4) -> bytes:
    delay = 3.0
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA})
            with urllib.request.urlopen(req, timeout=120) as resp:
                return resp.read()
        except Exception as exc:  # noqa: BLE001
            if attempt == retries - 1:
                raise
            print(f"  [retry {attempt+1}] {exc}", file=sys.stderr)
            time.sleep(delay)
            delay *= 2
    return b""


# Known historical workbook URLs (from a one-time crawl of the site's per-month
# article pages, whose numeric download IDs aren't otherwise discoverable). The
# light scrape below adds new months as PZPM posts them; this seed carries the
# back-history. (year, month) -> "download/<id>/<id>" path.
SEED_FILES = {
    (2023, 9): "/en/content/download/17654/224028",
    (2023, 10): "/en/content/download/17824/225570",
    (2024, 1): "/en/content/download/20410/250652",
    (2024, 3): "/en/content/download/20401/250574",
    (2024, 4): "/en/content/download/22059/265901",
    (2024, 5): "/en/content/download/19038/237712",
    (2024, 9): "/en/content/download/20431/250834",
    (2025, 3): "/en/content/download/22047/265797",
    (2025, 5): "/en/content/download/22056/265875",
    (2025, 9): "/en/content/download/22074/266031",
    (2025, 10): "/en/content/download/22077/266057",
    (2026, 1): "/en/content/download/18379/231085",
    (2026, 2): "/en/content/download/18361/230869",
    (2026, 3): "/en/content/download/18358/230833",
    (2026, 4): "/en/content/download/18382/231121",
}


def _seed_url(path: str, y: int, m: int) -> str:
    return f"{BASE}{path}/file/Press_release_car_registrations_{_FULL[m-1].title()}_{y}.xlsx"


def _get_text(path: str) -> str:
    try:
        return _get(path if path.startswith("http") else BASE + path).decode("utf-8", "replace")
    except Exception as exc:  # noqa: BLE001
        print(f"  [warn] fetch {path} failed: {exc}", file=sys.stderr)
        return ""


def _scan_xlsx(html: str, into: dict) -> None:
    for path, mon, yr in re.findall(
            r'(/en/content/download/\d+/\d+/file/Press_release_car_registrations_([A-Za-z]+)_(\d{4})\.xlsx)', html):
        i = MONTHS.get(mon.lower())
        if i:
            into[(int(yr), i)] = BASE + path


def discover_month_files() -> dict[tuple[int, int], str]:
    """Seed the known historical files, then a light scrape of the root + recent
    year pages to catch newly posted months (keeps the daily run fast)."""
    found: dict[tuple[int, int], str] = {
        (y, m): _seed_url(p, y, m) for (y, m), p in SEED_FILES.items()}
    root_html = _get_text(ROOT)
    _scan_xlsx(root_html, found)
    recent = {f"{ROOT}/Year-{y}" for y in (2025, 2026)}
    for yl in re.findall(r'"(' + re.escape(ROOT) + r'/Year-[0-9]+)"', root_html):
        recent.add(yl)
    for yl in sorted(recent):
        _scan_xlsx(_get_text(yl), found)
        time.sleep(0.1)
    return found


def parse_workbook(blob: bytes, file_year: int, month: int) -> list[tuple]:
    """Parse the Market (monthly) sheet, returning both the current month and the
    year-earlier month: [(year, month, country, rec, is_current)]."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb["Market (monthly)"] if "Market (monthly)" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[5:]:
        name = r[0]
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        if name.upper() in AGGREGATES or name[0].isdigit() or name.startswith("*"):
            continue

        def rec_for(colmap):
            def val(col):
                v = r[col] if col < len(r) else None
                try:
                    return int(round(float(v)))
                except (TypeError, ValueError):
                    return 0
            return {k: val(c) for k, c in colmap.items()}

        cur = rec_for(CUR_COLS)
        if cur["total"] > 0:
            out.append((file_year, month, name, cur, True))
        prior = rec_for(PRIOR_COLS)
        if prior["total"] > 0:
            out.append((file_year - 1, month, name, prior, False))
    return out


def load_existing() -> dict:
    data = {}
    if OUT_CSV.exists():
        with OUT_CSV.open(encoding="utf-8") as fh:
            for r in csv.DictReader(fh):
                data[(int(r["year"]), int(r["month"]), r["country"])] = r
    return data


FIELDS = ["year", "month", "country", "total", "bev", "phev", "hev", "other", "petrol", "diesel"]


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.parse_args()

    # (year, month, country) -> row dict, and whether it came from a current-year
    # reading (which takes precedence over a year-earlier reading of the same key).
    existing = {k: (v, True) for k, v in load_existing().items()}
    known_file_months = {(k[0], k[1]) for k in existing}

    available = discover_month_files()
    # Re-fetch the two most recent file-months (revisions) plus any not yet parsed.
    refresh = set(sorted(available)[-2:])
    new = sorted(k for k in available if k not in known_file_months or k in refresh)
    print(f"[acea] {len(available)} file(s) available; {len(new)} to fetch")

    for (fy, fm) in new:
        try:
            recs = parse_workbook(_get(available[(fy, fm)]), fy, fm)
        except Exception as exc:  # noqa: BLE001
            print(f"  {fy}-{fm:02d}: parse failed ({exc}), skipping", file=sys.stderr)
            continue
        n_cur = 0
        for (y, m, name, rec, is_current) in recs:
            key = (y, m, name)
            row = ({"year": y, "month": m, "country": name, **rec}, is_current)
            if is_current:
                existing[key] = row
                n_cur += 1
            elif key not in existing:      # don't overwrite a current reading
                existing[key] = row
        print(f"  file {fy}-{fm:02d}: {n_cur} countries (+ year-ago column)")
        time.sleep(0.3)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS)
        w.writeheader()
        for key in sorted(existing, key=lambda k: (k[0], k[1], k[2])):
            row = existing[key][0]
            w.writerow({f: row[f] for f in FIELDS})
    months = sorted({(k[0], k[1]) for k in existing})
    print(f"[write] {OUT_CSV.relative_to(REPO_ROOT)} ({len(existing)} rows, "
          f"{len(months)} months {months[0][0]}-{months[0][1]:02d}..{months[-1][0]}-{months[-1][1]:02d})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
