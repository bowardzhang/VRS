#!/usr/bin/env python3
"""Parse KBA FZ 10.1 workbooks into a tidy dataset for analysis / the website.

Reads every ``data/Germany/fz10_<YYYY>_<MM>.xlsx`` workbook, extracts the
brand / model-series rows from the ``FZ 10.1`` sheet, and writes:

* ``data/Germany/processed/germany_registrations.csv`` — one tidy row per
  (month, brand, model, drivetrain) with the monthly count, year-to-date
  count and share.
* ``docs/data/germany.json`` — a compact summary consumed by the static site
  (monthly totals, brand rankings, fuel-mix and top models).

The reference month is taken from the file name, not from the German month
label inside the sheet, so the parser is language- and layout-stable.
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data" / "Germany"
PROCESSED_DIR = DATA_DIR / "processed"
SITE_DATA_DIR = REPO_ROOT / "docs" / "data"

# Supplementary monthly powertrain series (BEV / diesel / plug-in hybrid /
# petrol) that predates the full workbooks we hold locally. It lets the site
# draw a multi-year powertrain trend even for months whose FZ 10.1 workbook is
# not present in ``data/Germany``. Full workbooks always take precedence.
POWERTRAIN_CSV = DATA_DIR / "kba_monthly_powertrain.csv"

SHEET_NAME = "FZ 10.1"
FILE_RE = re.compile(r"fz10_(\d{4})_(\d{2})\.xlsx$", re.IGNORECASE)

# Column of the *monthly* value for each drivetrain group. The next two
# columns hold the year-to-date value and the share (%). See row 8/9 headers.
DRIVETRAINS: dict[str, int] = {
    "total": 4,
    "diesel": 7,
    "hybrid_incl_plugin": 10,
    "petrol_hybrid_incl_plugin": 13,
    "diesel_hybrid_incl_plugin": 16,
    "hybrid_excl_plugin": 19,
    "petrol_hybrid_excl_plugin": 22,
    "diesel_hybrid_excl_plugin": 25,
    "plugin_hybrid": 28,
    "petrol_plugin_hybrid": 31,
    "diesel_plugin_hybrid": 34,
    "bev": 37,
    "awd": 40,
    "cabriolet": 43,
}

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# Rows in column 2 that are aggregates rather than individual brands.
GRAND_TOTAL_KEY = "NEUZULASSUNGEN INSGESAMT"
OTHER_KEY = "SONSTIGE ZUSAMMEN"


def _num(value) -> float | None:
    """Coerce a KBA cell to a number; '-' / blanks / text become None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text in {"", "-", ".", "x", "X"}:
        return None
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


# The data sheet is usually named "FZ 10.1", but the exact label varies across
# publications ("FZ 10.1 ", "FZ10.1", "FZ 10.1_neu", …). Match it loosely and
# skip obvious cover/notes sheets.
_SHEET_RE = re.compile(r"fz\s*10[._ ]*1", re.IGNORECASE)
_SKIP_SHEET_RE = re.compile(r"deckblatt|impressum|inhalt|erläuterung|hinweise", re.IGNORECASE)


def _pick_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    """Return the FZ 10.1 data sheet, tolerating naming variations."""
    names = wb.sheetnames
    if SHEET_NAME in names:
        return wb[SHEET_NAME]
    for name in names:
        if _SHEET_RE.search(name):
            return wb[name]
    for name in names:
        if not _SKIP_SHEET_RE.search(name):
            return wb[name]
    return wb[names[0]]


def parse_workbook(path: Path) -> tuple[int, int, list[dict]]:
    """Return (year, month, rows) for one FZ 10.1 workbook."""
    m = FILE_RE.search(path.name)
    if not m:
        raise ValueError(f"Unexpected file name: {path.name}")
    year, month = int(m.group(1)), int(m.group(2))

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = _pick_sheet(wb)

    rows: list[dict] = []
    current_brand: str | None = None
    for row in ws.iter_rows(min_row=10, values_only=True):
        # Cells are 0-indexed here; workbook column N -> row[N-1].
        brand_cell = (row[1] or "").strip() if isinstance(row[1], str) else row[1]
        model_cell = (row[2] or "").strip() if isinstance(row[2], str) else row[2]
        total_month = _num(row[3])

        if not brand_cell and not model_cell and total_month is None:
            continue

        label = str(brand_cell or "").strip()
        upper = label.upper()

        # Aggregate rows.
        if upper == GRAND_TOTAL_KEY or upper == OTHER_KEY:
            record_brand, record_model, kind = label, "", "aggregate"
        elif upper.endswith("ZUSAMMEN"):
            # Brand subtotal row, e.g. "ALFA ROMEO ZUSAMMEN".
            record_brand = label[: -len("ZUSAMMEN")].strip()
            record_model, kind = "", "brand_total"
            current_brand = record_brand
        elif label:
            # New brand header with its first model in column 3.
            current_brand = label
            record_brand, record_model, kind = label, str(model_cell or "").strip(), "model"
        else:
            # Continuation model row for the current brand.
            record_brand = current_brand or ""
            record_model, kind = str(model_cell or "").strip(), "model"

        metrics: dict[str, dict] = {}
        for name, col in DRIVETRAINS.items():
            metrics[name] = {
                "month": _num(row[col - 1]),
                "ytd": _num(row[col]),
                "share": _num(row[col + 1]),
            }
        rows.append(
            {
                "brand": record_brand,
                "model": record_model,
                "kind": kind,
                "metrics": metrics,
            }
        )
    wb.close()
    return year, month, rows


def write_tidy_csv(all_rows: list[dict]) -> Path:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out = PROCESSED_DIR / "germany_registrations.csv"
    with out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            ["year", "month", "brand", "model", "row_type", "drivetrain",
             "count_month", "count_ytd", "share_pct"]
        )
        for r in all_rows:
            for name, mvals in r["metrics"].items():
                if mvals["month"] is None and mvals["ytd"] is None:
                    continue
                writer.writerow(
                    [r["year"], r["month"], r["brand"], r["model"], r["kind"],
                     name, mvals["month"], mvals["ytd"], mvals["share"]]
                )
    return out


def _n(x):
    """Render a numeric metric as int when whole, else float, else None."""
    if x is None:
        return None
    return int(x) if float(x).is_integer() else x


# Country of origin by marque (the car's nationality, i.e. where the brand is
# from — not the current owner group; VW-owned Škoda stays Czech, Geely-owned
# Volvo stays Swedish). Brands not listed fall into "Other".
BRAND_ORIGIN: dict[str, str] = {
    # Germany
    "VW": "Germany", "MERCEDES": "Germany", "BMW": "Germany", "AUDI": "Germany",
    "OPEL": "Germany", "PORSCHE": "Germany", "SMART": "Germany", "MAN": "Germany",
    "ALPINA": "Germany", "BORGWARD": "Germany",
    # Czechia / Spain (VW group marques, kept at marque origin)
    "SKODA": "Czechia", "SEAT": "Spain", "CUPRA": "Spain",
    # France
    "RENAULT": "France", "PEUGEOT": "France", "CITROEN": "France", "DS": "France",
    "ALPINE": "France", "DACIA": "France",
    # Italy
    "FIAT": "Italy", "ALFA ROMEO": "Italy", "FERRARI": "Italy", "MASERATI": "Italy",
    "LANCIA": "Italy", "ABARTH": "Italy", "IVECO": "Italy",
    # Sweden
    "VOLVO": "Sweden", "POLESTAR": "Sweden",
    # United Kingdom
    "MINI": "United Kingdom", "LAND ROVER": "United Kingdom", "JAGUAR": "United Kingdom",
    "BENTLEY": "United Kingdom", "ROLLS ROYCE": "United Kingdom", "LOTUS": "United Kingdom",
    "ASTON MARTIN": "United Kingdom", "MORGAN": "United Kingdom", "INEOS": "United Kingdom",
    "MG": "China", "MG ROEWE": "China",  # SAIC-owned; treated as China below
    # USA
    "FORD": "USA", "TESLA": "USA", "JEEP": "USA", "CADILLAC": "USA",
    "CHEVROLET": "USA", "LUCID": "USA", "FISKER": "USA",
    # Japan
    "TOYOTA": "Japan", "MAZDA": "Japan", "NISSAN": "Japan", "SUZUKI": "Japan",
    "MITSUBISHI": "Japan", "HONDA": "Japan", "LEXUS": "Japan", "SUBARU": "Japan",
    "INFINITI": "Japan",
    # South Korea
    "HYUNDAI": "South Korea", "KIA": "South Korea", "GENESIS": "South Korea",
    "SSANGYONG": "South Korea", "KGM": "South Korea",
    # China
    "BYD": "China", "LEAPMOTOR": "China", "GWM": "China", "XPENG": "China",
    "LYNK & CO": "China", "NIO": "China", "GEELY": "China", "JAECOO": "China",
    "ZEEKR": "China", "MAXUS": "China", "OMODA": "China", "DEEPAL": "China",
    "CHERY": "China", "AIWAYS": "China", "DONGFENG": "China", "HONGQI": "China",
    "SERES": "China", "WEY": "China", "ORA": "China", "DFSK": "China",
}

# Groups always shown in the origin chart when present (the rest fold to "Other").
PREFERRED_ORIGINS = ["Germany", "China", "USA"]
MAX_BRAND_SERIES = 8
MAX_ORIGIN_SERIES = 8


def _brand_month_totals(month: dict) -> dict[str, float]:
    """{brand: monthly total} from a workbook month's brand-subtotal rows."""
    out: dict[str, float] = {}
    for r in month["rows"]:
        if r["kind"] == "brand_total":
            val = r["metrics"]["total"]["month"]
            if val is not None:
                out[r["brand"].strip().upper()] = val
    return out


def build_dimension_series(months: list[dict]) -> dict:
    """Multi-month series for the brand and country-of-origin dimensions.

    Only the months backed by a full FZ 10.1 workbook carry brand detail, so
    these series span those months (the powertrain trend, which also uses the
    supplementary CSV, is handled separately).
    """
    detailed = sorted(months, key=lambda mo: (mo["year"], mo["month"]))
    if not detailed:
        return {"brand_trends": None, "origin_trends": None}

    labels = [f"{MONTH_NAMES[mo['month'] - 1]} {mo['year']}" for mo in detailed]
    per_month = [_brand_month_totals(mo) for mo in detailed]

    # ---- brand trends: the biggest brands over the whole period ----
    brand_total: dict[str, float] = {}
    for totals in per_month:
        for brand, val in totals.items():
            brand_total[brand] = brand_total.get(brand, 0) + val
    top_brands = [b for b, _ in sorted(
        brand_total.items(), key=lambda kv: kv[1], reverse=True)][:MAX_BRAND_SERIES]
    brand_series = [
        {"name": b, "values": [_n(m.get(b)) for m in per_month]}
        for b in top_brands
    ]

    # ---- origin trends: registrations aggregated by marque nationality ----
    origin_month: list[dict[str, float]] = []
    origin_total: dict[str, float] = {}
    for totals in per_month:
        agg: dict[str, float] = {}
        for brand, val in totals.items():
            origin = BRAND_ORIGIN.get(brand, "Other")
            agg[origin] = agg.get(origin, 0) + val
            origin_total[origin] = origin_total.get(origin, 0) + val
        origin_month.append(agg)

    ranked = [o for o, _ in sorted(
        origin_total.items(), key=lambda kv: kv[1], reverse=True) if o != "Other"]
    named = [o for o in PREFERRED_ORIGINS if o in origin_total]
    for o in ranked:
        if len(named) >= MAX_ORIGIN_SERIES:
            break
        if o not in named:
            named.append(o)
    # Stable, largest-first ordering for a readable stack.
    named.sort(key=lambda o: origin_total[o], reverse=True)

    origin_series = [
        {"name": o, "values": [_n(m.get(o)) for m in origin_month]}
        for o in named
    ]
    other_vals = [
        _n(sum(v for k, v in m.items() if k not in named)) for m in origin_month
    ]
    if any(v for v in other_vals):
        origin_series.append({"name": "Other", "values": other_vals})

    return {
        "brand_trends": {"labels": labels, "series": brand_series},
        "origin_trends": {"labels": labels, "series": origin_series},
    }


def merge_powertrain_history(monthly_summary: list[dict]) -> list[dict]:
    """Add trend-only months from the supplementary powertrain CSV.

    Months already produced from a full FZ 10.1 workbook are kept as-is (they
    carry brand / model detail); months present only in the CSV are appended as
    lightweight entries so the powertrain trend can span several years. Returns
    the combined list sorted chronologically.
    """
    if not POWERTRAIN_CSV.exists():
        return sorted(monthly_summary, key=lambda mo: (mo["year"], mo["month"]))

    covered = {(mo["year"], mo["month"]) for mo in monthly_summary}
    with POWERTRAIN_CSV.open(encoding="utf-8") as fh:
        for rec in csv.DictReader(fh):
            year, month = int(rec["year"]), int(rec["month"])
            if (year, month) in covered:
                continue

            def _int(key):
                val = rec.get(key)
                return int(val) if val not in (None, "") else None

            monthly_summary.append(
                {
                    "year": year,
                    "month": month,
                    "label": f"{MONTH_NAMES[month - 1]} {year}",
                    "total": None,
                    "total_ytd": None,
                    "bev": _int("bev"),
                    "diesel": _int("diesel"),
                    "plugin_hybrid": _int("plugin_hybrid"),
                    "petrol": _int("petrol"),
                    "hybrid_incl_plugin": None,
                    "top_brands": [],
                    "top_models": [],
                }
            )
    return sorted(monthly_summary, key=lambda mo: (mo["year"], mo["month"]))


def build_site_json(months: list[dict]) -> dict:
    """Assemble the compact summary the static site renders."""
    months.sort(key=lambda mo: (mo["year"], mo["month"]))
    monthly_summary = []
    for mo in months:
        rows = mo["rows"]
        total_row = next(
            (r for r in rows if r["brand"].upper() == GRAND_TOTAL_KEY), None
        )
        brand_totals = [r for r in rows if r["kind"] == "brand_total"]

        def brand_metric(r, key):
            return r["metrics"][key]["month"] or 0

        brands_sorted = sorted(
            brand_totals, key=lambda r: brand_metric(r, "total"), reverse=True
        )
        top_brands = [
            {
                "brand": r["brand"],
                "total": _n(r["metrics"]["total"]["month"]),
                "bev": _n(r["metrics"]["bev"]["month"]),
                "diesel": _n(r["metrics"]["diesel"]["month"]),
                "plugin_hybrid": _n(r["metrics"]["plugin_hybrid"]["month"]),
            }
            for r in brands_sorted[:15]
        ]

        models = [r for r in rows if r["kind"] == "model"]
        top_models = sorted(
            models, key=lambda r: r["metrics"]["total"]["month"] or 0, reverse=True
        )[:15]
        top_models_out = [
            {
                "brand": r["brand"],
                "model": r["model"],
                "total": _n(r["metrics"]["total"]["month"]),
                "bev": _n(r["metrics"]["bev"]["month"]),
            }
            for r in top_models
        ]

        grand = total_row["metrics"] if total_row else None

        # Pure-petrol registrations are not a column of their own in FZ 10.1;
        # derive them so the powertrain trend has a consistent petrol series:
        # petrol = total - diesel - (all hybrids incl. plug-in) - bev.
        petrol = None
        if grand:
            parts = [grand[k]["month"] for k in
                     ("total", "diesel", "hybrid_incl_plugin", "bev")]
            if all(p is not None for p in parts):
                petrol = parts[0] - parts[1] - parts[2] - parts[3]

        monthly_summary.append(
            {
                "year": mo["year"],
                "month": mo["month"],
                "label": f"{MONTH_NAMES[mo['month'] - 1]} {mo['year']}",
                "total": _n(grand["total"]["month"]) if grand else None,
                "total_ytd": _n(grand["total"]["ytd"]) if grand else None,
                "bev": _n(grand["bev"]["month"]) if grand else None,
                "diesel": _n(grand["diesel"]["month"]) if grand else None,
                "plugin_hybrid": _n(grand["plugin_hybrid"]["month"]) if grand else None,
                # Full (non-plug-in) hybrids — only available from workbooks.
                "hybrid": _n(grand["hybrid_excl_plugin"]["month"]) if grand else None,
                "petrol": _n(petrol),
                "hybrid_incl_plugin": _n(grand["hybrid_incl_plugin"]["month"]) if grand else None,
                "top_brands": top_brands,
                "top_models": top_models_out,
            }
        )

    monthly_summary = merge_powertrain_history(monthly_summary)

    # The latest fully-detailed month (with brand / model rankings) drives the
    # headline sections; trend-only months from the supplementary series never
    # become "latest".
    latest = next(
        (mo for mo in reversed(monthly_summary) if mo.get("top_brands")), None
    )
    dims = build_dimension_series(months)
    return {
        "country": "Germany",
        "source": "Kraftfahrt-Bundesamt (KBA), table FZ 10.1",
        "source_url": (
            "https://www.kba.de/DE/Statistik/Produktkatalog/produkte/"
            "Fahrzeuge/fz10/fz10_gentab.html"
        ),
        "metric": "New passenger-car registrations (Neuzulassungen)",
        "months": monthly_summary,
        "latest": latest,
        "brand_trends": dims["brand_trends"],
        "origin_trends": dims["origin_trends"],
    }


def main() -> int:
    files = sorted(DATA_DIR.glob("fz10_*.xlsx"))
    if not files:
        print(f"No workbooks found in {DATA_DIR}")
        return 1

    all_rows: list[dict] = []
    months: list[dict] = []
    for path in files:
        year, month, rows = parse_workbook(path)
        for r in rows:
            all_rows.append({"year": year, "month": month, **r})
        months.append({"year": year, "month": month, "rows": rows})
        print(f"[parsed] {path.name}: {len(rows)} rows")

    csv_path = write_tidy_csv(all_rows)
    print(f"[write]  {csv_path.relative_to(REPO_ROOT)} ({len(all_rows)} records)")

    site = build_site_json(months)
    SITE_DATA_DIR.mkdir(parents=True, exist_ok=True)
    json_path = SITE_DATA_DIR / "germany.json"
    json_path.write_text(json.dumps(site, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[write]  {json_path.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
